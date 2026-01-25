"""
Requirements
    - pip install openpyxl flask
"""

from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from flask import jsonify

app = Flask(__name__)
app.secret_key = "local-dev-key"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ATTENDANCE_XLSX = os.path.join(BASE_DIR, "attendance.xlsx")

ATTENDANCE_SHEET = "Attendance"
PEOPLE_SHEET = "People"


# ----------------------------
# Excel setup helpers
# ----------------------------
def ensure_workbook_exists():
    """
    Ensures attendance.xlsx exists and contains:
      - Sheet 'Attendance' with header: event_date, full_name, sex, timestamp
      - Sheet 'People' with header: full_name, sex
    """
    if not os.path.exists(ATTENDANCE_XLSX):
        wb = Workbook()

        # Default sheet -> rename to Attendance
        ws_att = wb.active
        ws_att.title = ATTENDANCE_SHEET
        ws_att.append(["event_date", "full_name", "sex", "timestamp"])

        # Create People sheet
        ws_people = wb.create_sheet(PEOPLE_SHEET)
        ws_people.append(["full_name", "sex"])

        wb.save(ATTENDANCE_XLSX)
        return

    # If file exists, ensure both sheets + headers exist
    wb = load_workbook(ATTENDANCE_XLSX)

    if ATTENDANCE_SHEET not in wb.sheetnames:
        ws_att = wb.create_sheet(ATTENDANCE_SHEET, 0)
        ws_att.append(["event_date", "full_name", "sex", "timestamp"])
    else:
        ws_att = wb[ATTENDANCE_SHEET]
        if ws_att.max_row < 1:
            ws_att.append(["event_date", "full_name", "sex", "timestamp"])

    if PEOPLE_SHEET not in wb.sheetnames:
        ws_people = wb.create_sheet(PEOPLE_SHEET)
        ws_people.append(["full_name", "sex"])
    else:
        ws_people = wb[PEOPLE_SHEET]
        if ws_people.max_row < 1:
            ws_people.append(["full_name", "sex"])

    wb.save(ATTENDANCE_XLSX)


def normalize_name(name: str) -> str:
    return " ".join((name or "").strip().split())


def normalize_sex(sex: str) -> str:
    return (sex or "").strip().upper()


# ----------------------------
# People sheet (data source)
# ----------------------------
def load_people():
    """
    Returns list of dicts: [{"full_name": "...", "sex": "M/F"}, ...]
    Reads from attendance.xlsx -> sheet 'People'
    """
    ensure_workbook_exists()
    wb = load_workbook(ATTENDANCE_XLSX)
    ws = wb[PEOPLE_SHEET]

    people = []
    # Expect header row at row 1: full_name, sex
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = normalize_name(row[0] if len(row) > 0 else "")
        sex = normalize_sex(row[1] if len(row) > 1 else "")

        if full_name and sex in {"M", "F"}:
            people.append({"full_name": full_name, "sex": sex})

    people.sort(key=lambda p: p["full_name"].lower())
    return people


def get_sex_by_name(full_name: str) -> str:
    target = normalize_name(full_name).lower()
    for p in load_people():
        if p["full_name"].lower() == target:
            return p["sex"]
    return ""


def add_person_to_people_sheet(full_name: str, sex: str) -> bool:
    """
    Adds a new person to People sheet if not already present (case-insensitive).
    Returns True if added, False if duplicate/invalid.
    """
    ensure_workbook_exists()

    new_name = normalize_name(full_name)
    sex = normalize_sex(sex)

    if not new_name or sex not in {"M", "F"}:
        return False

    existing_lower = {p["full_name"].lower() for p in load_people()}
    if new_name.lower() in existing_lower:
        return False

    wb = load_workbook(ATTENDANCE_XLSX)
    ws = wb[PEOPLE_SHEET]
    ws.append([new_name, sex])
    wb.save(ATTENDANCE_XLSX)
    return True


# ----------------------------
# Attendance sheet (logging)
# ----------------------------
def append_attendance(event_date: str, full_name: str):
    ensure_workbook_exists()

    full_name = normalize_name(full_name)
    if not full_name:
        return

    sex = get_sex_by_name(full_name)

    wb = load_workbook(ATTENDANCE_XLSX)
    ws = wb[ATTENDANCE_SHEET]

    ws.append([
        event_date,
        full_name,
        sex,
        datetime.now().isoformat(timespec="seconds"),
    ])

    wb.save(ATTENDANCE_XLSX)


# ----------------------------
# Routes
# ----------------------------
@app.route("/", methods=["GET"])
def index():
    people = load_people()
    return render_template("index.html", people=people)


@app.route("/submit", methods=["POST"])
def submit():
    event_date = (request.form.get("event_date") or "").strip()
    names_raw = (request.form.get("full_names") or "").strip()

    if not event_date or not names_raw:
        flash("Please select at least one name and event date.")
        return redirect(url_for("index"))

    names = [normalize_name(n) for n in names_raw.split(",") if normalize_name(n)]

    present = get_attendance_names_for_date(event_date)

    added_count = 0
    for full_name in names:
        if full_name.lower() in present:
            continue
        append_attendance(event_date, full_name)
        added_count += 1

    flash(f"Recorded attendance for {added_count} people. (Ignored duplicates: {len(names)-added_count})")
    return redirect(url_for("index"))


@app.route("/add_person", methods=["POST"])
def add_person():
    full_name = request.form.get("new_full_name", "")
    sex = request.form.get("sex", "")
    event_date = (request.form.get("event_date") or "").strip()

    new_name = normalize_name(full_name)

    if len(new_name) < 2:
        flash("Name is too short.")
        return redirect(url_for("index"))

    added = add_person_to_people_sheet(new_name, sex)

    if added:
        # Mark attendance immediately if date exists
        if event_date:
            append_attendance(event_date, new_name)
            flash(f"Added & recorded attendance: {new_name}")
        else:
            flash(f"Added new person: {new_name}")
    else:
        flash("Person already exists or invalid sex.")

    return redirect(url_for("index"))

def get_attendance_names_for_date(event_date: str) -> set[str]:
    ensure_workbook_exists()
    wb = load_workbook(ATTENDANCE_XLSX, data_only=True)
    ws = wb[ATTENDANCE_SHEET]

    present = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = (row[0] or "")
        name = normalize_name(row[1] or "")
        if d == event_date and name:
            present.add(name.lower())
    return present


@app.get("/attendance_status")
def attendance_status():
    event_date = (request.args.get("event_date") or "").strip()
    if not event_date:
        return jsonify({"event_date": "", "present": []})

    present = sorted(list(get_attendance_names_for_date(event_date)))
    return jsonify({"event_date": event_date, "present": present})


if __name__ == "__main__":
    # Make sure workbook/sheets exist before first request
    ensure_workbook_exists()
    app.run(host="127.0.0.1", port=5000, debug=True)